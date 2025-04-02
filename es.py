import streamlit as st
import base64
import time
import os
import datetime
import pandas as pd
from streamlit_extras.grid import grid
from streamlit_extras.stylable_container import stylable_container
from PIL import Image
import folium
import plotly.graph_objects as go
from geopy.geocoders import Nominatim
from openpyxl import load_workbook
import geopandas as gpd
from folium import Choropleth
from streamlit_folium import folium_static
import json
import re
from functions import *
import hashlib




st.set_page_config(
    page_title="Blood Donation Campaign Dashboard",
    page_icon="🩸",
    layout="wide",
    initial_sidebar_state="expanded"
)

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# Fonction pour vérifier les identifiants
def check_credentials(username, password):
    users = {
        "outsider": hash_password("pass_outsider"),
    }
    if username in users and users[username] == hash_password(password):
        return True
    return False

# Fonction pour créer le fichier de utilisateurs (à exécuter une fois pour initialiser)
def create_users_file():
    if not os.path.exists("users.csv"):
        users_data = {
            "username": ["user1", "user2", "admin"],
            "password": [
                hash_password("password1"),
                hash_password("password2"),
                hash_password("admin123")
            ]
        }
        pd.DataFrame(users_data).to_csv("users.csv", index=False)
        st.success("Fichier d'utilisateurs créé avec succès!")

# Configuration de session state pour garder l'état de connexion
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

# Interface de connexion
def login_form():
    st.title("Connexion")
    
    with st.form("login_form"):
        username = st.text_input("Nom d'utilisateur")
        password = st.text_input("Mot de passe", type="password")
        submit = st.form_submit_button("Se connecter")
        
        if submit:
            if check_credentials(username, password):
                st.session_state["authenticated"] = True
                st.session_state["username"] = username
                st.success("Connexion réussie! valider a nouveau pour confirmer la connexion")
                st.rerun()
            else:
                st.error("Nom d'utilisateur ou mot de passe incorrect")

#_____________________________________
def main_app():
    if st.sidebar.button("Déconnexion"):
        st.session_state["authenticated"] = False
        st.rerun()

    if 'language' not in st.session_state:
        st.session_state.language = "English"

    # Chargement des traductions
    with open('langage.json', 'r', encoding='utf-8') as json_file:
        TRANSLATIONS = json.load(json_file)

    def get_modification_time(path = "donnees.xlsx"):
        return os.path.getmtime(path)

    #_____________________Function_____________________________________

    # Cache pour les shapefiles avec simplification de géométrie
    @st.cache_resource
    def load_shapefile(shapefile_path, simplify_tolerance=None):
        gdf = gpd.read_file(shapefile_path)
        if simplify_tolerance is not None:
            gdf.geometry = gdf.geometry.simplify(simplify_tolerance)
        return gdf


    # Fonction pour charger les données avec la date de modification comme hash
    @st.cache_resource
    def load_data1():
        df = pd.read_excel('last.xlsx')
        #df = pd.read_excel('Challenge dataset.xlsx')
        #df = pd.read_excel('DataChallengeFinale1.xlsx')
        return df
    def load_data3():
        df = pd.read_excel('last.xlsx', sheet_name="2020")
        return df
    @st.cache_resource
    def load_data2(get_modification_time):
        df = pd.read_excel('donnees.xlsx')
        return df

    def get_combined_data():
        return pd.concat([load_data1(), load_data2(get_modification_time())], axis=0)


    #________________________


    def get_hierarchical_data():
        """Prétraiter les données avec les relations hiérarchiques entre niveaux administratifs"""
        # Charger les shapefiles avec les relations hiérarchiques
        geo_data_1 = load_shapefile("gadm41_CMR_1.shp", simplify_tolerance=0.01)
        geo_data_2 = load_shapefile("gadm41_CMR_2.shp", simplify_tolerance=0.01)
        geo_data_3 = load_shapefile("gadm41_CMR_3.shp", simplify_tolerance=0.01)
        
        # Créer un dictionnaire de correspondance entre arrondissements et départements/régions
        admin_hierarchy = pd.DataFrame({
            'NAME_3': geo_data_3['NAME_3'],
            'NAME_2': geo_data_3['NAME_2'], 
            'NAME_1': geo_data_3['NAME_1']
        })
        
        # Fusionner les données des volontaires avec la hiérarchie administrative
        combined_data = get_combined_data()
        enriched_data = combined_data.merge(
            admin_hierarchy, 
            left_on='Arrondissement_de_résidence_', 
            right_on='NAME_3', 
            how='left'
        )
        
        return enriched_data



    @st.cache_resource
    def create_card(content, key, cell_height="200px", cell_width="160px"):
        """Create a styled card with custom dimensions"""
        with stylable_container(
            key=key,
            css_styles=f"""
                {{
                    min-height: {cell_height};
                    height: {cell_height};
                    width: {cell_width};
                    max-width: {cell_width};
                    border: 1px solid #c0c0c0;
                    border-radius: 10px;
                    margin: 0px;  # Small margin for spacing
                    padding: 0px;
                    display: flex;
                    flex-direction: column;
                    align-items: center;
                    justify-content: center;
                    background-color: #FFFEBA;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.2);
                    transition: all 0.2s ease;
                    overflow: hidden;
                    text-overflow: ellipsis;
                    color: green;
                }}
                
                .card-title {{
                        font-weight: bold;
                        margin: 0px;
                        padding: 0px;
                        font-size: 1em;
                        text-align: center;
                        color: #8a2be2;  # Light purple color
                    }}
                
                .card-value {{
                    font-size:1.2em;
                    text-align: center;
                }}
                
            """
        ):
            st.markdown(content, unsafe_allow_html=True)


    def create_metric_card(column, title, plot_function, height="250px",width="100%"):
        cell_width = width
        key = f"{title.lower().replace(' ', '_')}_card"
        
        with column:
            # Create content for the card heading
            header_content = f"""
            <div class="card-title">{title}</div>
            """
            
            # Use the stylable container with reusable styling
            with stylable_container(
                key=key,
                css_styles=f"""
                    {{
                        min-height: {height};
                        height: auto;
                        width: {cell_width};
                        max-width: {cell_width};
                        border: 1px solid black;
                        border-radius: 10px;
                        background-color: #f8f9fa;
                        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.8);
                        transition: all 0.2s ease;
                        overflow: hidden;
                        text-overflow: ellipsis;
                        color: #8a2be2;
                    }}
                    
                    .card-title {{
                        font-weight: bold;
                        margin-top: 0px;
                        padding: 0px;
                        font-size: 1.2em;
                        text-align: center;
                    }}
                    
                    .card-value {{
                        color: black;
                        font-size: 1em;
                        font-weight: bold;
                        text-align: center;
                    }}
                    
                """
            ):
                # Render the header content
                st.markdown(header_content, unsafe_allow_html=True)
                
                # Get the plot from the provided function
                fig = plot_function()
            




    def get_preprocessed_data(level):
        """Prétraiter les données pour différents niveaux administratifs"""
        enriched_data = get_hierarchical_data()
        
        if level == 1:  # Région
            data_counts = enriched_data['NAME_1'].value_counts().reset_index()
            data_counts.columns = ['NAME_1', 'Nb']
            return data_counts
            
        elif level == 2:  # Département
            data_counts = enriched_data['NAME_2'].value_counts().reset_index()
            data_counts.columns = ['NAME_2', 'Nb']
            return data_counts
            
        elif level == 3:  # Arrondissement
            data_counts = enriched_data['Arrondissement_de_résidence_'].value_counts().reset_index()
            data_counts.columns = ['Arrondissement_de_résidence_', 'Nb']
            return data_counts
        
        return None



    # Helper function to format card content
    @st.cache_resource
    def format_card_content(title, value):
        
        return f"""
            <div class='card-title' style='margin: 0px; padding: 0px;'>{title}</div>
            <div class='card-value' style='margin: 0px; padding: 0px;'>{value}</div>
        """
    #<div class='card-delta' style='margin: 0px; padding: 0px;'>{delta_text}</div>______________________________________________________________

    if "donors" not in st.session_state:
        st.session_state.donors = pd.DataFrame(columns=[
        'id', 'age', 'Date_remplissage','Date_naiss', 'niveau_detude', 'genre', 'taille', 'poids',
        'situation_matrimoniale', 'profession', 'arrondissement_residence',
        'quartier_residence', 'nationalite', 'religion', 'deja_donne_sang',
        'date_dernier_don', 'taux_dhemoglobine', 'eligibilite_au_don',
        'est_sous_anti_biotherapie', 'taux_dhemoglobine_bas', 'date_dernier_don_3_mois',
        'ist_recente', 'ddr_incorrecte', 'allaitement', 'accouchement_6mois',
        'interruption_grossesse', 'enceinte', 'antecedent_transfusion',
        'porteur_hiv_hbs_hcv', 'opere', 'drepanocytaire', 'diabetique',
        'hypertendus', 'asthmatiques', 'cardiaque', 'tatoue', 'scarifie',
        'autres_raisons'
    ])
            
            # Add a sample donor record
    # Import the challenge dataset
    #challenge_data = pd.read_excel('Challenge dataset.xlsx')
    # Apply custom CSS for the sidebar styling
    @st.cache_resource
    def get_text(key):
        if 'language' not in st.session_state:
            return key
        if 'get_text' not in st.session_state:
            return key
        return st.session_state.get_text(key)
    st.markdown('''
    <style>
        /* Sidebar styles */
        .sidebar , .sidebar-content {
            display: flex;
            flex-direction: column;
            align-items: flex-start;
            justify-content: flex-start;
        }

        .sidebar, .styled-header {
            font-size: 18px;
            font-weight: bold;
            color: WHITE !important;
            text-align: center;
            margin-top: 20px;
            background-color: rgba(25, 0, 0, 0.4);
            padding: 10px;
            border-radius: 6px;
            box-shadow: 10 4px 8px rgba(0, 0, 0, 0.1);
        }

        /* Styling for the sidebar links */
        .sidebar, .sidebar-nav {
            display: flex;
            flex-direction: column;
            align-items: flex-start;
            width: 100%;
        }

        .sidebar-nav a {
            text-decoration: none;
            padding: 12px;
            color: blue;
            font-weight: bold;
            margin-top: 10px;
            width: 100%;
            text-align: left;
            border-radius: 5px;
            background-color: #f1f1f1;
        }

        .sidebar-nav a:hover {
            background-color: #e1e1e1;
        }
        .stButton{
            color: white;  /* Nouvelle couleur du texte lors du survol */
            font-weight: bold;  /* Nouveau poids de la police lors du survol */
            
        }
        .stButton>button:hover {
            background-color:rgb(90, 6, 6);  /* Nouvelle couleur de fond lors du survol */
            color: white;  /* Nouvelle couleur du texte lors du survol */
            font-weight: bold;
            transition: background-color 0.3s 
        }

        }
    </style>
    ''', unsafe_allow_html=True)
    st.sidebar.image('logo.jpeg', width=200)

    # Define the menu items

    def get_text(key):
        if key not in TRANSLATIONS[st.session_state.language]:
            return key  # Return key if translation doesn't exist
        return TRANSLATIONS[st.session_state.language][key]

    menu_items = {
        "Home": f"🏠 {get_text('Home')}",
        "Donations": f"💉 {get_text('Donations')}",
        "Cartography": f"📍 {get_text('Cartography')}",
        "Dataset Insights": f"🔄 {get_text('Dataset Insights')}",
        "Eligibility and Profile": f"🩺 {get_text('Eligibility and Profile')}",
        "Campaign Insights": f"📊 {get_text('Campaign Insights')}",
        "Fidélisation": f"📊 {get_text('Fidélisation')}",
        "Options": f"⚙️ {get_text('Options')}",
        "About": f"ℹ️ {get_text('About')}"
    }

    # Create a selectbox for the menu
    selected_item = st.sidebar.selectbox(
        " ",
        options=list(menu_items.keys()), 
        format_func=lambda x: menu_items[x],
        index=0,  # Premier élément sélectionné par défaut
        key="main_menu",
        help="Sélectionnez une option du menu"
    )

    if selected_item =="Fidélisation":
        st.markdown("""
            <style>
            .block-container {
                padding-top: 0rem;
                padding-bottom: 0rem;
                padding-left:4rem;
                padding-right: 4rem;
                margin-top: 0px;
            }
                
                /* Remove extra padding around header */
                header {
                    margin-bottom: 0rem !important;
                    padding-bottom: 0rem !important;
                }
            </style>
            """, unsafe_allow_html=True)
        
        combined_data = get_combined_data()
        row1_cols = st.columns(3)
        with row1_cols[1]:
            all_arrondissement = combined_data['Arrondissement_de_résidence_'].unique()
            selected_arrondissement = st.sidebar.multiselect(
                get_text("Districts"),
                all_arrondissement
            )
            if not selected_arrondissement:
                selected_arrondissement = all_arrondissement
        with row1_cols[2]:
            all_case = combined_data['ÉLIGIBILITÉ_AU_DON.'].unique()
            selected_case = st.sidebar.multiselect(
                "Eligible",
                all_case # Limiter par défaut pour améliorer les performances
            )
            if not selected_case:
                selected_case = all_case
        # Filtrer les données avec tous les filtres
        filtered_data = combined_data[
            (combined_data['Arrondissement_de_résidence_'].isin(selected_arrondissement)) &
            (combined_data['ÉLIGIBILITÉ_AU_DON.'].isin(selected_case))]
        
    # Étape 1 : Filtrer les donneurs éligibles
        df = load_data1()
        df_eligible = df[df['ÉLIGIBILITÉ_AU_DON.'] == 'Eligible']

        # Identifier les donneurs récurrents et non récurrents
        df_eligible_recurrent = df_eligible[df_eligible['A-t-il (elle) déjà donné le sang'] == 'Oui']
        df_eligible_non_recurrent = df_eligible[df_eligible['A-t-il (elle) déjà donné le sang'] == 'Non']

        # Liste des colonnes démographiques à analyser
        demographic_columns = ['Classe_Age', 'Genre_', "Niveau_d'etude", 'Religion_Catégorie', 
                            'Situation_Matrimoniale_(SM)', 'categories', 'Arrondissement_de_résidence_']

        # Fonction pour générer un graphique interactif avec les top 4 catégories
        def plot_top4_demographic(data_recurrent, data_non_recurrent, column, title_prefix, comparison=False, orientation='v'):
            count_recurrent = data_recurrent[column].value_counts()
            if comparison:
                count_non_recurrent = data_non_recurrent[column].value_counts()
                
                all_categories = pd.concat([count_recurrent, count_non_recurrent], axis=1, sort=False)
                all_categories.columns = ['Récurrents', 'Non Récurrents']
                all_categories.fillna(0, inplace=True)
                
                all_categories['Total'] = all_categories['Récurrents'] + all_categories['Non Récurrents']
                top4_categories = all_categories.sort_values('Total', ascending=False).head(4).index
                
                count_recurrent = count_recurrent[count_recurrent.index.isin(top4_categories)]
                count_non_recurrent = count_non_recurrent[count_non_recurrent.index.isin(top4_categories)]
                
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=count_recurrent.index if orientation == 'v' else count_recurrent.values,
                    y=count_recurrent.values if orientation == 'v' else count_recurrent.index,
                    name='Récurrents (Oui)',
                    marker_color='#00cc96',
                    text=count_recurrent.values,
                    textposition='auto',
                    orientation='h' if orientation == 'h' else 'v'
                ))
                fig.add_trace(go.Bar(
                    x=count_non_recurrent.index if orientation == 'v' else count_non_recurrent.values,
                    y=count_non_recurrent.values if orientation == 'v' else count_non_recurrent.index,
                    name='Non Récurrents (Non)',
                    marker_color='#ff5733',
                    text=count_non_recurrent.values,
                    textposition='auto',
                    orientation='h' if orientation == 'h' else 'v'
                ))
            else:
                top4_categories = count_recurrent.head(4).index
                count_recurrent = count_recurrent[count_recurrent.index.isin(top4_categories)]
                
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=count_recurrent.index if orientation == 'v' else count_recurrent.values,
                    y=count_recurrent.values if orientation == 'v' else count_recurrent.index,
                    name='Récurrents',
                    marker_color='#00cc96',
                    text=count_recurrent.values,
                    textposition='auto',
                    orientation='h' if orientation == 'h' else 'v'
                ))
            
            fig.update_layout(
                title=f"{title_prefix} par {column} (Top 4)",
                xaxis_title=column if orientation == 'v' else 'Nombre de donneurs',
                yaxis_title='Nombre de donneurs' if orientation == 'v' else column,
                xaxis=dict(tickangle=45),
                legend=dict(title='Statut de récurrence', orientation='h', yanchor='bottom', y=-0.3, xanchor='center', x=0.5),
                plot_bgcolor='white',
                width=800,
                height=600,
                barmode='group' if comparison else 'stack'
            )
            
            return fig

        # Initialize Streamlit
        st.title(get_text("Blood Donor Analysis"))
        st.write("Select a demographic variable to visualize the distribution of recurrent and non-recurrent donors.")

        # Sélection des paramètres par l'utilisateur
        selected_column = st.selectbox(get_text("Choose a demographic variable:"), demographic_columns)
        orientation = st.radio(get_text("Graph orientation:"), ['Vertical', 'Horizontal'])
        comparison = st.checkbox(get_text("Compare recurrent and non-recurrent donors?"), value=False)

        # Convertir en format utilisable par la fonction
        orientation_value = 'v' if orientation == 'Verticale' else 'h'

        # Générer le graphique
        fig = plot_top4_demographic(df_eligible_recurrent, df_eligible_non_recurrent, selected_column, "Analyse des donneurs", comparison, orientation_value)

        # Afficher le graphique dans Streamlit
        st.plotly_chart(fig)

        ###########################"
        df_temp_non_eligible = df[df['ÉLIGIBILITÉ_AU_DON.'] == 'Temporairement Non-eligible']

        # Séparer les hommes et les femmes
        df_temp_men = df_temp_non_eligible[df_temp_non_eligible['Genre_'] == 'Homme']
        df_temp_women = df_temp_non_eligible[df_temp_non_eligible['Genre_'] == 'Femme']

        # Définir une palette de couleurs
        color_men = '#1f77b4'
        color_women = '#ff7f0e'

        # Fonction pour générer un graphique Plotly
        def plot_top4_demographic(data_men, data_women, column, title_prefix, orientation='v'):
            count_men = data_men[column].value_counts()
            count_women = data_women[column].value_counts()
            
            all_categories = pd.concat([count_men, count_women], axis=1, sort=False).fillna(0)
            all_categories.columns = ['Hommes', 'Femmes']
            all_categories['Total'] = all_categories['Hommes'] + all_categories['Femmes']
            top4_categories = all_categories.sort_values('Total', ascending=False).head(4).index
            
            count_men = count_men[count_men.index.isin(top4_categories)]
            count_women = count_women[count_women.index.isin(top4_categories)]
            
            fig = go.Figure()
            
            if orientation == 'v':
                fig.add_trace(go.Bar(x=count_men.index, y=count_men.values, name='Hommes', marker_color=color_men))
                fig.add_trace(go.Bar(x=count_women.index, y=count_women.values, name='Femmes', marker_color=color_women))
                fig.update_layout(xaxis_title=column, yaxis_title='Donor number')
            else:
                fig.add_trace(go.Bar(y=count_men.index, x=count_men.values, name='Hommes', marker_color=color_men, orientation='h'))
                fig.add_trace(go.Bar(y=count_women.index, x=count_women.values, name='Femmes', marker_color=color_women, orientation='h'))
                fig.update_layout(xaxis_title='Donor', yaxis_title=column)
            
            return fig
            # Interface Streamlit
        st.title(get_text("Analysis of Temporarily Non-Eligible Donors"))

        demographic_columns = {
            'Classe_Age': get_text('Age Group'),
            'categories': get_text('Professional Categories'),
            'Arrondissement_de_résidence_': get_text('District of Residence'),
            'Raison_indisponibilité_fusionnée': get_text('Reasons for Ineligibility')
        }

        selected_column = st.selectbox(get_text("select a catégory"), list(demographic_columns.keys()), format_func=lambda x: demographic_columns[x])

        graph_orientation = 'h' if selected_column in ['categories', 'Arrondissement_de_résidence_', 'Raison_indisponibilité_fusionnée'] else 'v'

        st.plotly_chart(plot_top4_demographic(df_temp_men, df_temp_women, selected_column, "Profil des donneurs", orientation=graph_orientation))
    #_________________________________________________________________
        def plot_bar_reasons(df, eligibility_type='Temporairement Non-eligible', gender=None):
            filtered_df = df[df['ÉLIGIBILITÉ_AU_DON.'] == eligibility_type]
            if gender:
                filtered_df = filtered_df[filtered_df['Genre_'] == gender]
            
            reasons_list = filtered_df['Raison_indisponibilité_fusionnée'].dropna().str.split(';').explode().str.strip()
            reason_counts = Counter(reasons_list)
            
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=list(reason_counts.keys()),
                y=list(reason_counts.values()),
                marker_color='#FF6347',
                text=[str(val) for val in reason_counts.values()],
                textposition='auto',
                textfont=dict(size=16, color='white', family='Arial Black')
            ))
            
            title = f"{get_text('Reasons for ineligibility')} ({eligibility_type})"
            if gender:
                title += f" - {gender}"
            
            fig.update_layout(
                title=title,
                xaxis_title=get_text("Ineligibility reasons"),
                yaxis_title=get_text("Number of occurrences"),
                template='plotly_white',
                bargap=0.2,
                xaxis=dict(tickangle=-45)
            )
            
            return fig

        def plot_frequencies_by_category(df, category_col, gender_col):
            count_data = df.groupby([category_col, gender_col]).size().unstack(fill_value=0)
            categories = count_data.index
            
            fig = go.Figure()
            for gender in count_data.columns:
                fig.add_trace(go.Bar(
                    x=categories, 
                    y=count_data[gender], 
                    name=gender,
                ))
            
            fig.update_layout(
                title=f"{get_text('Distribution of')} {category_col} {get_text('by gender')}",
                xaxis_title=category_col,
                yaxis_title=get_text("Number of people"),
                barmode='group'
            )
            return fig

        def plot_hemoglobin_box(df, eligibility_col='ÉLIGIBILITÉ_AU_DON.', hemoglobin_col='Taux d’hémoglobine'):
            if not all(col in df.columns for col in [eligibility_col, hemoglobin_col]):
                st.error(get_text("The required columns are missing from the dataset."))
                return None
            
            df[hemoglobin_col] = pd.to_numeric(df[hemoglobin_col], errors='coerce')
            df_clean = df.dropna(subset=[hemoglobin_col])
            
            fig = go.Figure()
            colors = ['#FF4040', '#FF8C00', '#FFB6C1']
            
            for i, status in enumerate(df_clean[eligibility_col].unique()):
                status_data = df_clean[df_clean[eligibility_col] == status][hemoglobin_col]
                fig.add_trace(go.Box(
                    y=status_data,
                    name=status,
                    marker_color=colors[i % len(colors)],
                    boxpoints='all', jitter=0.3, pointpos=-1.8, line_width=2
                ))
            
            fig.update_layout(
                title=get_text("Hemoglobin level distribution by eligibility status"),
                xaxis_title=get_text("Eligibility status"),
                yaxis_title=get_text("Hemoglobin level (g/dL)"),
                template='plotly_white',
                showlegend=True,
                height=600
            )
            
            fig.add_hline(y=12.5, line_dash="dash", line_color="red", annotation_text=get_text("Min threshold (F)"), annotation_position="top right")
            fig.add_hline(y=13.0, line_dash="dash", line_color="blue", annotation_text=get_text("Min threshold (H)"), annotation_position="top right")
            
            return fig

        st.title(get_text("Donor Ineligibility Analysis"))
        col = st.columns(2)
        with col[0]:
            type_eligibility = st.selectbox(get_text("Choose an eligibility type"), ['Temporairement Non-eligible', 'Définitivement non-eligible'])
        with col[1]:
            gender = st.selectbox(get_text("Filter by gender (optional)"), [get_text('All'), 'Homme', 'Femme'])
        if gender == get_text('All'):
            gender = None
        st.plotly_chart(plot_bar_reasons(df, type_eligibility, gender))

        demographic_columns = {
            'Classe_Age': get_text("Age range"),
            'categories': get_text("Professional categories"),
            'Arrondissement_de_résidence_': get_text("Residence district"),
            'Raison_indisponibilité_fusionnée': get_text("Ineligibility reasons")
        }
        selected_column = st.selectbox(get_text("Select a category to analyze"), list(demographic_columns.keys()), format_func=lambda x: demographic_columns[x])
        graph_orientation = 'h' if selected_column in ['categories', 'Arrondissement_de_résidence_', 'Raison_indisponibilité_fusionnée'] else 'v'
        st.plotly_chart(plot_frequencies_by_category(df_temp_non_eligible, selected_column, 'Genre_'))
        st.title(get_text("Hemoglobin Level Analysis"))
        st.plotly_chart(plot_hemoglobin_box(df))
        #_______________________________________________________

    if selected_item =="Campaign Insights":
        st.markdown("""
    <style>
        .card-title {
            font-weight: bold;
            font-size: 1.2em;
            text-align: center;
            padding: 10px 0;
            color: #8a2be2;
        }
        
        /* Make plots use full width of their containers */
        .js-plotly-plot, .plotly, .plot-container {
            width: 100% !important;
        }
        
        /* Remove default padding from columns to maximize space */
        div[data-testid="column"] {
            padding: 0 !important;
            margin: 0 !important;
        }
        
        /* Style for the streamlit columns inside the container */
        div[data-testid="stHorizontalBlock"] {
            width: 100% !important;
            gap: 5px;
        }
    </style>
    """, unsafe_allow_html=True)
        st.markdown(
        """
        <style>
            div[data-baseweb="select"] > div {
                min-height: 10px !important;  /* Adjust height */
            }
        </style>
        """,
        unsafe_allow_html=True
    )
        st.markdown("""
            <style>
            .block-container {
                padding-top: 0rem;
                padding-bottom: 0rem;
                padding-left:4rem;
                padding-right: 4rem;
                margin-top: 0px;
            }
                
                /* Remove extra padding around header */
                header {
                    margin-bottom: 0rem !important;
                    padding-bottom: 0rem !important;
                }
            </style>
            """, unsafe_allow_html=True)
        combined_data = get_combined_data()
        row1_cols = st.columns(3)
        with row1_cols[1]:
            all_arrondissement = combined_data['Arrondissement_de_résidence_'].unique()
            selected_arrondissement = st.sidebar.multiselect(
                get_text("Districts"),
                all_arrondissement
            )
            if not selected_arrondissement:
                selected_arrondissement = all_arrondissement
        with row1_cols[2]:
            all_case = combined_data['ÉLIGIBILITÉ_AU_DON.'].unique()
            selected_case = st.sidebar.multiselect(
                "Eligible",
                all_case # Limiter par défaut pour améliorer les performances
            )
            if not selected_case:
                selected_case = all_case
        # Filtrer les données avec tous les filtres
        filtered_data = combined_data[
            (combined_data['Arrondissement_de_résidence_'].isin(selected_arrondissement)) &
            (combined_data['ÉLIGIBILITÉ_AU_DON.'].isin(selected_case))]
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("Welcome to the  campaign Insight Section")}
        </div> """,  unsafe_allow_html=True)
        filtered_data['Date de remplissage de la fiche'] = pd.to_datetime(df['Date de remplissage de la fiche'])
        m= filtered_data['Date de remplissage de la fiche'].dt.month
        month_counts = m.value_counts().sort_index()
        categories = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        data = {"Nombre d'occurences": month_counts.reindex(range(1, 13), fill_value=0).tolist()}

        filtered_data1 = combined_data[(combined_data['ÉLIGIBILITÉ_AU_DON.']=="Définitivement non-eligible")]
        filtered_data2 = combined_data[(combined_data['ÉLIGIBILITÉ_AU_DON.']=="Temporairement Non-eligible")]

        row1_cos = st.columns(1)
        create_metric_card(row1_cos[0],title="Number of Donors By Months",plot_function=lambda: plot_radar_chart(data, categories),width="100%") 
        
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("")}</div> """,  unsafe_allow_html=True)
        row1_cos = st.columns(1)
        create_metric_card(row1_cos[0],title=get_text("Evolution of Number of Donors"),plot_function=lambda: number_line(filtered_data),width="100%") 
        
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("")}</div> """,  unsafe_allow_html=True)
        row1_cos = st.columns(1)
        create_metric_card(row1_cos[0],title=get_text("Evolution of Donations"),plot_function=lambda: heatmap(filtered_data),width="100%") 
        
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("")}</div> """,  unsafe_allow_html=True)
        row1_cos = st.columns(1)
        create_metric_card(row1_cos[0],title=get_text("Evolution of Donation"),plot_function=lambda: jour(load_data1(),height="400px"),width="100%") 

        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("")}</div> """,  unsafe_allow_html=True)
        row1_cos = st.columns(2)
        create_metric_card(row1_cos[0],title=get_text("Non Eligibility Purposes"),plot_function=lambda: generate_wordcloud_and_barchart(filtered_data1),width="100%") 
        create_metric_card(row1_cos[1],title=get_text("Non Availability purposes"),plot_function=lambda:generate_wordcloud_and_barchart(filtered_data2),width="100%") 

    if selected_item =="Eligibility and Profile" : 
        st.markdown("""
    <style>
        .card-title {
            font-weight: bold;
            font-size: 1.2em;
            text-align: center;
            padding: 10px 0;
            color: #8a2be2;
        }
        
        /* Make plots use full width of their containers */
        .js-plotly-plot, .plotly, .plot-container {
            width: 100% !important;
        }
        
        /* Remove default padding from columns to maximize space */
        div[data-testid="column"] {
            padding: 0 !important;
            margin: 0 !important;
        }
        
        /* Style for the streamlit columns inside the container */
        div[data-testid="stHorizontalBlock"] {
            width: 100% !important;
            gap: 5px;
        }
    </style>
    """, unsafe_allow_html=True)
        st.markdown(
        """
        <style>
            div[data-baseweb="select"] > div {
                min-height: 10px !important;  /* Adjust height */
            }
        </style>
        """,
        unsafe_allow_html=True
    )
        combined_data = get_combined_data()
        row1_cols = st.columns(3)
        with row1_cols[1]:
            all_arrondissement = combined_data['Arrondissement_de_résidence_'].unique()
            selected_arrondissement = st.sidebar.multiselect(
                get_text("Districts"),
                all_arrondissement
            )
            if not selected_arrondissement:
                selected_arrondissement = all_arrondissement
        with row1_cols[2]:
            all_case = combined_data['ÉLIGIBILITÉ_AU_DON.'].unique()
            selected_case = st.sidebar.multiselect(
                "Eligible",
                all_case # Limiter par défaut pour améliorer les performances
            )
            if not selected_case:
                selected_case = all_case
        # Filtrer les données avec tous les filtres
        filtered_data = combined_data[
            (combined_data['Arrondissement_de_résidence_'].isin(selected_arrondissement)) &
            (combined_data['ÉLIGIBILITÉ_AU_DON.'].isin(selected_case))]
        st.markdown("""
            <style>
            .block-container {
                padding-top: 0rem;
                padding-bottom: 0rem;
                padding-left: 1rem;
                padding-right: 1rem;
                margin-top: 0px;
            }
                
                /* Remove extra padding around header */
                header {
                    margin-bottom: 0rem !important;
                    padding-bottom: 0rem !important;
                }
            </style>
            """, unsafe_allow_html=True)
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("Welcome to the  Eligibility and Profile Section")}
        </div> """,  unsafe_allow_html=True)
        row1_cos = st.columns(1)
        create_metric_card(row1_cos[0],title=get_text("Eligibility Profile"),plot_function=lambda: circle(filtered_data),width="100%") 
        
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("")}</div> """,  unsafe_allow_html=True)
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("")}<br/></div> """,  unsafe_allow_html=True)
        if st.session_state.language == "English"  : 
            create_metric_card(row1_cos[0],title=get_text("Ideal Donor"),plot_function=lambda: display_ideal_chart_e() , width="100%") 
        else :  
            create_metric_card(row1_cos[0],title="Ideal Donor",plot_function=lambda: display_ideal_chart_f() , width="100%") 
        
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("")}</div> """,  unsafe_allow_html=True)
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("")}</div> """,  unsafe_allow_html=True)
        create_metric_card(row1_cos[0],title=get_text("Eligibility In Douala"),plot_function=lambda: three(filtered_data),width="100%") 


    if selected_item == "Dataset Insights":
        st.markdown("""
    <style>
        .card-title {
            font-weight: bold;
            font-size: 1.2em;
            text-align: center;
            padding: 10px 0;
            color: #8a2be2;
        }
        
        /* Make plots use full width of their containers */
        .js-plotly-plot, .plotly, .plot-container {
            width: 100% !important;
        }
        
        /* Remove default padding from columns to maximize space */
        div[data-testid="column"] {
            padding: 0 !important;
            margin: 0 !important;
        }
        
        /* Style for the streamlit columns inside the container */
        div[data-testid="stHorizontalBlock"] {
            width: 100% !important;
            gap: 5px;
        }
    </style>
    """, unsafe_allow_html=True)
        st.markdown(
        """
        <style>
            div[data-baseweb="select"] > div {
                min-height: 10px !important;  /* Adjust height */
            }
        </style>
        """,
        unsafe_allow_html=True
    )
        st.markdown("""
            <style>
            .block-container {
                padding-top: 0rem;
                padding-bottom: 0rem;
                padding-left: 1rem;
                padding-right: 1rem;
                margin-top: 0px;
            }
                
                /* Remove extra padding around header */
                header {
                    margin-bottom: 0rem !important;
                    padding-bottom: 0rem !important;
                }
            </style>
            """, unsafe_allow_html=True)
        combined_data = get_combined_data()
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("Welcome to the  social demographics insights")}
        </div> """,  unsafe_allow_html=True)
        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;"> </div> """,  unsafe_allow_html=True)
        all_arrondissement = combined_data['Arrondissement_de_résidence_'].unique()
        selected_arrondissement = st.sidebar.multiselect(
            get_text("Districts"),
            all_arrondissement
        )
        if not selected_arrondissement:
            selected_arrondissement = all_arrondissement

        all_case = combined_data['ÉLIGIBILITÉ_AU_DON.'].unique()
        selected_case = st.sidebar.multiselect(
            "Eligible",
            all_case # Limiter par défaut pour améliorer les performances
        )
        if not selected_case:
            selected_case = all_case
        # Filtrer les données avec tous les filtres
        filtered_data = combined_data[
            (combined_data['Arrondissement_de_résidence_'].isin(selected_arrondissement)) &
            (combined_data['ÉLIGIBILITÉ_AU_DON.'].isin(selected_case))]

        cols = st.columns(2)
        create_metric_card(cols[0], title=get_text("Marital Status"), plot_function=lambda: render_frequency_pieh(filtered_data["Situation_Matrimoniale_(SM)"], legend_top="70%", legend_left="15%"), width="100%")
        create_metric_card(cols[1], title=get_text("Level"), plot_function=lambda: render_frequency_pieh(filtered_data["Niveau_d'etude"]), width="100%")
        cols = st.columns(2)
        create_metric_card(cols[0], title=get_text("Sector"), plot_function=lambda: render_frequency_pie(filtered_data["Secteur"], legend_left="2%", legend_top="2%"), width="100%")
        create_metric_card(cols[1], title=get_text("Gender"), plot_function=lambda: render_frequency_pieh(filtered_data["Genre_"], legend_top="80%", legend_left="65%"), width="100%")
        
        cols = st.columns(2)
        create_metric_card(cols[0], title=get_text("Profession"), plot_function=lambda: render_frequency_pie(filtered_data["categories"], legend_left="2%", legend_top="2%"), width="100%")
        create_metric_card(cols[1], title=get_text("Eligibility Status"), plot_function=lambda: render_frequency_pieh(filtered_data["ÉLIGIBILITÉ_AU_DON."], legend_top="90%", legend_left="15%"), width="100%")

        cols = st.columns(2)
        option = compute_age_distribution(filtered_data['Age'])
        create_metric_card(cols[0], title=get_text('Age distribution'), plot_function=lambda: st_echarts(options=option, height=400, width="100%"), width="100%")
        create_metric_card(cols[1], title=get_text('Population Pyramid'), plot_function=lambda: plot_age_pyramid(filtered_data, height=400), width="100%")

    if selected_item == "Options":
        st.markdown(
            """
            <h1 style='color: red; text-align: center; margin-top: 0;'>Options</h1>
            """,
            unsafe_allow_html=True
        )
        st.markdown(get_text('### Theme Personalization'))
        # Initialize theme in session state
        if 'theme' not in st.session_state:
            st.session_state.theme = "Light"

        # Define themes with fonts and colors
        THEMES = {
            "Light": {
                "primaryColor": "#F50307",
                "backgroundColor": "#FBFBFB",
                "secondaryBackgroundColor": "#EC8282",
                "textColor": "#000000",
                "font": "sans serif"
            },
            "Dark": {
                "primaryColor": "#f50307",
                "backgroundColor": "#0E1117",
                "secondaryBackgroundColor": "#2D2A2A",
                "textColor": "#fff4f4",
                "font": "sans serif"
            },
            "Blue": {
                "primaryColor": "#087CEF",
                "backgroundColor": "#FFFFFF",
                "secondaryBackgroundColor": "#7FD6AC",
                "textColor": "black",
                "font": "sans serif"
            }
        }   
        @st.cache_resource 
        def apply_theme(theme_choice):
            # Get the selected theme
            selected_theme = THEMES[theme_choice]

            # Update session state
            st.session_state.theme = theme_choice

            # Ensure the .streamlit directory exists
            config_dir = os.path.join(os.path.expanduser("~"), ".streamlit")
            os.makedirs(config_dir, exist_ok=True)

            # Write the selected theme to the config.toml file
            config_path = os.path.join(config_dir, "config.toml")
            with open(config_path, "w") as config_file:
                config_file.write(f"""
                [theme]
                primaryColor = "{selected_theme['primaryColor']}"
                backgroundColor = "{selected_theme['backgroundColor']}"
                secondaryBackgroundColor = "{selected_theme['secondaryBackgroundColor']}"
                textColor = "{selected_theme['textColor']}"
                font = "{selected_theme['font']}"
                """)
            # Show success message
            st.success(f"Theme will be changed to {theme_choice}! Click Again to confirm. ")
    

        # Theme selection dropdown
        theme = st.selectbox(
            get_text("Choose your theme"), 
            list(THEMES.keys()),
            index=list(THEMES.keys()).index(st.session_state.theme),
            key="widget_theme"
        )

        # Get selected theme colors
        selected_theme = THEMES[st.session_state.theme]

        # Apply button with dynamic text color
        apply_button = st.markdown(
            f"""
            <style>
                div.stButton > button {{
                    color: {selected_theme['textColor']} !important;
                    border-radius: 5px;
                    padding: 10px 20px;
                    font-size: 16px;
                }}
            </style>
            """,
            unsafe_allow_html=True
        )

        # Add confirmation button to apply theme
        if st.button(get_text("Apply Theme"), key="apply_theme_button"):
            apply_theme(st.session_state.widget_theme)
    #__________________________________
        # Initialize language if not set
        if 'language' not in st.session_state:
            st.session_state.language = list(TRANSLATIONS.keys())[0]  # Default to first language

        # Function to apply language
    
        def apply_language():
            st.session_state.language = st.session_state.widget_language

            
        # Language selection with callback
        st.selectbox(
            get_text("Choose your language"), 
            list(TRANSLATIONS.keys()),
            index=list(TRANSLATIONS.keys()).index(st.session_state.language),
            key="widget_language"
        )

        if st.button(get_text("Apply Language"), key="apply_language_button"):
            st.success(f"Language will be changed to {st.session_state.widget_language  }! Click Again to confirm. ")
            apply_language()
                    
    if selected_item == "Home":
        n=0.811111
    
        col = st.columns(1)

        with col[0]: 
            with stylable_container(
                key='842',
                css_styles="""
                    {border: 1px solid #c0c0c0;
                    border-radius: 10px;
                    margin: 0px;  # Small margin for spacing
                    top : 7px;
                    padding: 10px ; 
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.2);
                    transition: all 0.2s ease;
                    overflow: hidden;
                    text-overflow: ellipsis;
                    color: green;
                    }
                    
                """
            ):
                st.markdown(
                    f"""
                    <div class="card-title" style="text-align: center; font-size: 29px; font-weight: bold; color: red;">
                        {get_text("Welcome to Our Blood Donation Campaign Dashboard")}
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            
        st.markdown("""
            <style>
            .block-container {
                padding-top: 2rem;
                padding-bottom: 0rem;
                padding-left: 2rem;
                padding-right: 2rem;
                margin-top: 0px;
            }
                
                /* Remove extra padding around header */
                header {
                    margin-bottom: 0rem !important;
                    padding-bottom: 0rem !important;
                }
            </style>
            """, unsafe_allow_html=True)
        row1_cols = st.columns([1,1,2,1.7,2,2,2])
        with row1_cols[0]:
            
            Volonteers = load_data1().shape[0] +load_data2(get_modification_time()).shape[0]
            temp_content = format_card_content(get_text("Volonteers"), Volonteers)
            create_card(temp_content, key="temperature", cell_height="90px", cell_width="105%")

        with row1_cols[1]:
            df1 = load_data1()
            df2 = load_data2(get_modification_time())
            eligible = df1[df1["ÉLIGIBILITÉ_AU_DON."] == "Eligible"].shape[0] + df2[df2["ÉLIGIBILITÉ_AU_DON."] == "Eligible"].shape[0]
            wind_content = format_card_content(get_text("Eligible"), eligible)
            create_card(wind_content, key="wind", cell_height="90px", cell_width="105%")

        with row1_cols[2]:
            T_eligible = df1[df1["ÉLIGIBILITÉ_AU_DON."] == "Temporairement Non-eligible"].shape[0] + df2[df2["ÉLIGIBILITÉ_AU_DON."] == "Temporairement Non-eligible"].shape[0]
            wind_content = format_card_content(get_text("Temporarily Non-eligible"), T_eligible)
            create_card(wind_content, key="humidity", cell_height="90px", cell_width="103%")

        with row1_cols[3]:
            T_eligible = df1[df1["ÉLIGIBILITÉ_AU_DON."] == "Définitivement non-eligible"].shape[0] + df2[df2["ÉLIGIBILITÉ_AU_DON."] == "Définitivement non-eligible"].shape[0]
            wind_content = format_card_content(get_text("Definitely Non-eligible"), T_eligible)
            create_card(wind_content, key="card", cell_height="90px", cell_width="103%")

        with row1_cols[6]:
            df =  get_combined_data()
            nomb = df['Si oui preciser la date du dernier don'].count()
            wind = format_card_content(get_text("Have Ever Donated"), nomb)
            create_card(wind, key="donated", cell_height="90px", cell_width="103%")

        with row1_cols[4]:
            df= load_data1()[(load_data1()['ÉLIGIBILITÉ_AU_DON.']=="Définitivement non-eligible")]
            df = df.dropna(subset=["Raison_indisponibilité_fusionnée"])
            df['Raison_indisponibilité_fusionnée'] = df["Raison_indisponibilité_fusionnée"].astype(str)
            df['Raison_indisponibilité_fusionnée'] = df["Raison_indisponibilité_fusionnée"].str.split(';')
            df_exploded = df.explode("Raison_indisponibilité_fusionnée")
            df_exploded = df_exploded.dropna(subset=["Raison_indisponibilité_fusionnée"])
            element_frequent = df_exploded["Raison_indisponibilité_fusionnée"].mode()[0]
            wind = format_card_content(get_text("Most Ilegibility Purpose"), element_frequent)
            create_card(wind, key="available", cell_height="90px", cell_width="103%")

        with row1_cols[5]:
            df= load_data1()[(load_data1()['ÉLIGIBILITÉ_AU_DON.']=="Temporairement Non-eligible")]
            df = df.dropna(subset=["Raison_indisponibilité_fusionnée"])
            df['Raison_indisponibilité_fusionnée'] = df["Raison_indisponibilité_fusionnée"].astype(str)
            df['Raison_indisponibilité_fusionnée'] = df["Raison_indisponibilité_fusionnée"].str.split(';')
            df_exploded = df.explode("Raison_indisponibilité_fusionnée")
            df_exploded = df_exploded.dropna(subset=["Raison_indisponibilité_fusionnée"])
            element_frequent = df_exploded["Raison_indisponibilité_fusionnée"].mode()[0]
            wind = format_card_content(get_text("Most Non Availability Purpose"), element_frequent)
            create_card(wind, key="eligible", cell_height="90px", cell_width="103%")

        gdf = load_shapefile("gadm41_CMR_0.shp")
        data_df_3 = get_preprocessed_data(3)
        data_df_2 = get_preprocessed_data(2)

        row=st.columns(1)
        with row[0]:
            with stylable_container(
                    key='944',
                    css_styles=f"""
                        {{
                        width : 100%;
                        border: 1px solid #c0c0c0;
                        border-radius: 10px;
                        flex-direction: column;
                        background-color: #f8f9fa;
                        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.8);
                        }}
                        
                        .card-title {{
                                font-weight: bold;
                                margin: 0px;
                                padding: 0px;
                                font-size: 1em;
                                text-align: center;
                                color: #8a2be2;  # Light purple color
                            }}
                    """ ):
                    st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 22px; font-weight: bold; color:black">{get_text("Keys Rate")}</div> """,  unsafe_allow_html=True)       
                    row = st.columns([1,1,1])
                    with row[0]:
                        option_jauge = {
                        "tooltip": {
                            "formatter": '{a} <br/>{b} : {c}%'
                        },
                        "series": [
                            {
                                "name": "Progression",
                                "type": "gauge",
                                "startAngle": 180,
                                "endAngle": 0,
                                "radius": "90%",
                                "itemStyle": {
                                    "color": "red",
                                    "shadowColor": "rgba(0,138,255,0.45)",
                                    "shadowBlur": 10,
                                    "shadowOffsetX": 2,
                                    "shadowOffsetY": 2
                                },
                                "progress": {
                                    "show": True,
                                    "roundCap": True,
                                    "width": 10
                                },
                                "pointer": {
                                    "length": "60%",
                                    "width": 3,
                                    "offsetCenter": [0, "5%"]
                                },
                                "detail": {
                                    "valueAnimation": True,
                                    "formatter": "{value}%",
                                    "backgroundColor": "red",
                                    "width": "100%",
                                    "lineHeight": 25,
                                    "height": 15,
                                    "borderRadius": 188,
                                    "offsetCenter": [0, "40%"],
                                    "fontSize": 12
                                },
                                "data": [{
                                    "value": round(100*eligible/Volonteers, 1),  # Example value
                                    "name": get_text("Eligibility Rate"),  # Label for the value"Eligibility Rate"
                                }]
                            }
                        ]
                    }
                        st_echarts(options=option_jauge, key="1")
                    with row[1]:
                        option_jauge = {
                        "tooltip": {
                            "formatter": '{a} <br/>{b} : {c}%'
                        },
                        "series": [
                            {
                                "name": "Progression",
                                "type": "gauge",
                                "startAngle": 180,
                                "endAngle": 0,
                                "radius": "90%",
                                "itemStyle": {
                                    "color": "red",
                                    "shadowColor": "rgba(0,138,255,0.45)",
                                    "shadowBlur": 10,
                                    "shadowOffsetX": 2,
                                    "shadowOffsetY": 2
                                },
                                "progress": {
                                    "show": True,
                                    "roundCap": True,
                                    "width": 10
                                },
                                "pointer": {
                                    "length": "60%",
                                    "width": 3,
                                    "offsetCenter": [0, "5%"]
                                },
                                "detail": {
                                    "valueAnimation": True,
                                    "formatter": "{value}%",
                                    "backgroundColor": "red",
                                    "width": "100%",
                                    "lineHeight": 25,
                                    "height": 15,
                                    "borderRadius": 188,
                                    "offsetCenter": [0, "40%"],
                                    "fontSize": 12
                                },
                                "data": [{
                                    "value": round(100*T_eligible/Volonteers, 1),  # Example value
                                    "name": get_text("Temporarily Eligibility Rate"), #"Temporarily Eligibility Rate"
                                }]
                            }
                        ]
                    }
                        st_echarts(options=option_jauge, key="2")
                    with row[2]:
                        option_jauge = {
                        "tooltip": {
                            "formatter": '{a} <br/>{b} : {c}%'
                        },
                        "series": [
                            {
                                "name": "Progression",
                                "type": "gauge",
                                "startAngle": 180,
                                "endAngle": 0,
                                "radius": "90%",
                                "itemStyle": {
                                    "color": "red",
                                    "shadowColor": "rgba(0,138,255,0.45)",
                                    "shadowBlur": 10,
                                    "shadowOffsetX": 2,
                                    "shadowOffsetY": 2
                                },
                                "progress": {
                                    "show": True,
                                    "roundCap": True,
                                    "width": 10
                                },
                                "pointer": {
                                    "length": "60%",
                                    "width": 3,
                                    "offsetCenter": [0, "5%"]
                                },
                                "detail": {
                                    "valueAnimation": True,
                                    "formatter": "{value}%",
                                    "backgroundColor": "red",
                                    "width": "100%",
                                    "lineHeight": 25,
                                    "height": 15,
                                    "borderRadius": 188,
                                    "offsetCenter": [0, "40%"],
                                    "fontSize": 12
                                },
                                "data": [{
                                    "value": round(100*(Volonteers-eligible-T_eligible)/Volonteers, 1),  # Example value
                                    "name": get_text("Ilegibility Rate"), #"Ilegibility Rate"
                                }]
                            }
                        ]
                    }
                        st_echarts(options=option_jauge, key="3")
                    

        row = st.columns([1.1,1.2])
        with row[0] : 
            st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 20px; font-weight: bold; color:black">{get_text("Donnor Insights")}</div> """,  unsafe_allow_html=True)       
            with stylable_container(
                key='1116',
                css_styles=f"""
                    {{
                    width : 100%;
                    border: 1px solid #c0c0c0;
                    border-radius: 10px;
                    flex-direction: column;
                    background-color: #f8f9fa;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.8);
                    }}
                    
                    .card-title {{
                            font-weight: bold;
                            margin: 0px;
                            padding: 0px;
                            font-size: 1em;
                            text-align: center;
                            color: #8a2be2;  # Light purple color
                        }}
                """ ):
                    row_1 = st.columns(2)
                    st.write("")
                    with row_1[0] :
                        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 20px; font-weight: bold; color:red">{get_text("Blood distribution")}</div> """,  unsafe_allow_html=True)       
                        pie_data = count_frequencies(load_data3()["Groupe Sanguin ABO / Rhesus "])
                        render_frequency_pie2(pie_data)
                    with row_1[1] :
                        st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 20px; font-weight: bold; color:red">{get_text("Donation Type")}</div> """,  unsafe_allow_html=True)       
                        pie_data = count_frequencies(load_data3()["Type de donation"])
                        render_frequency_pie2(pie_data)
                    st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 20px; font-weight: bold; color:red">{get_text("Donor Profile")}</div> """,  unsafe_allow_html=True)       
                    if st.session_state.language == "English":
                        display_ideal_chart_e() 
                    else : display_ideal_chart_f()
            st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 20px; font-weight: bold; color:black">{get_text("Social and Demographics stats")}</div> """,  unsafe_allow_html=True)       
            with stylable_container(
                key='1152',
                css_styles=f"""
                    {{width : 100%;
                    
        
                        border: 1px solid #c0c0c0;
                    border-radius: 10px;
                    margin: 0px;  # Small margin for spacing
                    padding: 0px;
                    display: flex;
                    flex-direction: column;
                    background-color: #f8f9fa;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.811111);
                    transition: all 0.2s ease;
                    overflow: hidden;
                    text-overflow: ellipsis;
                    color: green;
                    }}
                    
                    .card-title {{
                            font-weight: bold;
                            margin: 0px;
                            padding: 0px;
                            font-size: 1em;
                            text-align: center;
                            color: #8a2be2;  # Light purple color
                        }}
                """ ):
                row_1 = st.columns([1,1.4])
                with row_1[0] :
                    st.markdown(f'''<div style='text-align: center;'>{get_text("Level")}</div>''', unsafe_allow_html=True)
                    pie_data = count_frequencies(load_data1()["Niveau_d'etude"])
                    render_frequency_pie2(pie_data)
                    st.markdown(f'''<div style='text-align: center;'>{get_text("Marital Status")}</div>''', unsafe_allow_html=True)
                    pie_data = count_frequencies(load_data1()["Situation_Matrimoniale_(SM)"])
                    render_frequency_pie2(pie_data)
                with row_1[1] :
                    st.markdown(f'''<div style='text-align: center;'>{get_text("Gender")}</div>''', unsafe_allow_html=True)
                    pie_data = count_frequencies(load_data1()["Genre_"])
                    render_frequency_pie2(pie_data)
                    st.markdown(f'''<div style='text-align: center;'>{get_text("Eligibility")}</div>''', unsafe_allow_html=True)
                    pie_data = count_frequencies(load_data1()["ÉLIGIBILITÉ_AU_DON."])
                    render_frequency_pie2(pie_data)



        with row[1] : 
            st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 20px; font-weight: bold; color:black">{get_text("Cartography Summary")}</div> """,  unsafe_allow_html=True)       
            with stylable_container(
                key='1201',
                css_styles=f"""
                    {{width : 110%;
                        border: 1px solid #c0c0c0;
                    border-radius: 10px;
                    margin: 0px;  # Small margin for spacing
                    padding: 0px;
                    display: flex;
                    flex-direction: column;
                    background-color: #f8f9fa;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.811111);
                    transition: all 0.2s ease;
                    overflow: hidden;
                    text-overflow: ellipsis;
                    color: green;
                    }}
                    
                    .card-title {{
                            font-weight: bold;
                            margin: 0px;
                            padding: 0px;
                            font-size: 1em;
                            text-align: center;
                            color: #8a2be2;  # Light purple color
                        }}
                """ ):
                    st.write("")
                    row_1 = st.columns(2)
                    with row_1[1] :
                        st.dataframe(data_df_3.head(5), use_container_width=True)
                        st.dataframe(data_df_2.head(8), use_container_width=True)
                    with row_1[0] :
                        gdf = load_shapefile("gadm41_CMR_0.shp")
                        m = folium.Map(location=[10.87, 13.52], zoom_start=5, control_scale=True, tiles="CartoDB positron", height=901)
                        folium.GeoJson(
                            gdf.to_crs(epsg=4326).__geo_interface__,
                            name="Cameroun",
                            style_function=lambda x: {'fillColor': 'transparent', 'color': 'black', 'weight': 2, 'fillOpacity': 0.0}
                        ).add_to(m)
                        m.fit_bounds([[6.2, 9.5], [0, 16.75]])

                        combined_data = get_hierarchical_data()
                        geo_data_3 = load_shapefile("gadm41_CMR_3.shp", simplify_tolerance=0.01)

                        geo_data_3['centroid_lat'] = geo_data_3['geometry'].centroid.y
                        geo_data_3['centroid_lon'] = geo_data_3['geometry'].centroid.x

                        centroid_dict = dict(zip(geo_data_3['NAME_3'], zip(geo_data_3['centroid_lat'], geo_data_3['centroid_lon'])))

                        geo_data_3 = geo_data_3.merge(combined_data, left_on='NAME_3', right_on='Arrondissement_de_résidence_')
                        mask_missing = combined_data['latitude'].isna() | combined_data['longitude'].isna()
                        missing_indices = combined_data[mask_missing].index

                        for idx in missing_indices:
                            arrondissement = combined_data.loc[idx, 'Arrondissement_de_résidence_']
                            if arrondissement in centroid_dict:
                                combined_data.loc[idx, 'latitude'] = centroid_dict[arrondissement][0]
                                combined_data.loc[idx, 'longitude'] = centroid_dict[arrondissement][1]

                        combined_data = combined_data.dropna(subset=['latitude', 'longitude'])
                        quartier_counts = combined_data['Quartier_de_Résidence_'].value_counts().reset_index()
                        quartier_counts.columns = ['Quartier_de_Résidence_', 'count']

                        quartier_locations = combined_data.groupby('Quartier_de_Résidence_').agg({
                            'latitude': 'first',
                            'longitude': 'first'
                        }).reset_index()

                        quartier_data = quartier_counts.merge(quartier_locations, on='Quartier_de_Résidence_')

                        min_count = quartier_data['count'].min() if not quartier_data.empty else 1
                        max_count = quartier_data['count'].max() if not quartier_data.empty else 1
                        for _, row in quartier_data.iterrows():
                            quartier = row["Quartier_de_Résidence_"]
                            count = row["count"]

                            radius = 3 + 4 * ((count - min_count) / max(max_count - min_count, 1)) * 10

                            folium.CircleMarker(
                                location=[row["latitude"], row["longitude"]],
                                radius=radius,
                                tooltip=f"{quartier}: {count} {get_text('Volonteers')}",
                                color='red',
                                fill=True,
                                fill_color='red'
                            ).add_to(m)

                        folium.LayerControl().add_to(m)

                        folium_static(m, width=400, height=500)

            st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 20px; font-weight: bold; color:black">{get_text("Campaign efficacity")}</div> """,  unsafe_allow_html=True)       
            with stylable_container(
                key='1294',
                css_styles=f"""
                    {{width : 95%;
                    
                    left : 5%;
                        border: 1px solid #c0c0c0;
                    border-radius: 10px;
                    margin: 0px;  # Small margin for spacing
                    padding: 0px;
                    display: flex;
                    flex-direction: column;
                    background-color: #f8f9fa;
                    box-shadow: 0 4px 6px rgba(0, 0, 0, 0.811111);
                    transition: all 0.2s ease;
                    overflow: hidden;
                    text-overflow: ellipsis;
                    color: green;
                    }}
                    
                    .card-title {{
                            font-weight: bold;
                            margin: 0px;
                            padding: 0px;
                            font-size: 1em;
                            text-align: center;
                            color: #8a2be2;  # Light purple color
                        }}
                """ ):
                    st.write("")
                    df = load_data1()
                    df['Date de remplissage de la fiche'] = pd.to_datetime(df['Date de remplissage de la fiche'])
                    m= df['Date de remplissage de la fiche'].dt.month
                    month_counts = m.value_counts().sort_index()
                    categories = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
                    data = {"Nombre d'occurences": month_counts.reindex(range(1, 13), fill_value=0).tolist()}
                    jour(load_data1(),height="400px")
                    number_line(load_data1())  

        
    elif selected_item == "Donations":
        with st.sidebar:
            donations_button = st.sidebar.button("Registration", use_container_width=True)
            dataset_button = st.sidebar.button("Dataset", use_container_width=True)
            
    # Donor Registration page
        if not dataset_button:
            geolocator = Nominatim(user_agent="geoapi")
            st.markdown(f"""<div class="card-title" style="text-align: center; font-size: 24px; font-weight: bold; color: #FF5A5A;">
        {get_text("REGISTRATION")}
        </div> """,  unsafe_allow_html=True)

            with st.form("donor_registration_form"):
                col1, col2 = st.columns(2)
                last_donor_number = 0
                # Liste des pays d'Afrique centrale
                pays_afrique_centrale = ["Cameroun", "République centrafricaine", "Tchad", "République du Congo", "République démocratique du Congo", "Gabon", "Guinée équatoriale", "São Tomé-et-Príncipe", "Autres"]
                if not st.session_state.donors.empty:
                    last_donor_number = int(st.session_state.donors['id'].str.extract(r'(\d+)').max())
                elif not load_data2(get_modification_time()).empty:
                    last_donor_number = int(load_data2(get_modification_time())['ID'].str.extract(r'(\d+)').max())
                else:
                    last_donor_number = 0
                with col1:
                    import datetime
                    donor_number = st.number_input(get_text("Registration Number"), min_value=last_donor_number + 1, value=last_donor_number + 1, step=1)
                    ID = f"DONOR_{donor_number}"
                    Date_remplissage = st.date_input(get_text("Filling Date"))
                    Date_naiss = st.date_input(get_text("BirthDate"), min_value=datetime.date(1960, 1, 1))
                    st.write(f"Âge : {(Date_remplissage - Date_naiss).days //365} ans")
                    Niveau_detude = st.selectbox(get_text("Education Level"), ["Primary", "Secondary", "High School", "Bachelor's", "Master's", "PhD", "Other"])
                    Genre_ = st.radio(get_text("Gender"), ["Homme", "Femme"])
                    Taille_ = st.number_input(get_text("Height"), min_value=100, max_value=220, step=1)
                    Poids = st.number_input(get_text("Weight"), min_value=60, max_value=200, step=1)
                    Profession_ = st.text_input("Profession")  
                    
                                    
                with col2:
                    A_deja_donne_le_sang_ = st.radio(get_text("Has already donated blood"), ["Yes", "No"]) 
                    geo_data_3 = gpd.read_file("gadm41_CMR_3.shp")
                    districts = geo_data_3["NAME_3"].unique()
                    Arrondissement_de_residence_ = st.selectbox(
                        get_text ("District of Residence"),
                        options=districts,
                        index=0,  # Sélectionner le premier élément par défaut
                        help="Sélectionnez votre district de résidence"
                    )
                            
                    Quartier_de_Residence_ = st.text_input( get_text("residential neighborhood"))

                    # Search button for location
                    if Quartier_de_Residence_:
                        location = geolocator.geocode(f"{Quartier_de_Residence_}, Cameroun")
                        if location:
                            latitude = location.latitude 
                            longitude = location.longitude
                            Quartier_de_Residence_ = geolocator.geocode(f"{Quartier_de_Residence_}, Cameroun")
                            st.success(f"📍 {Quartier_de_Residence_} {get_text('Localised')} → Latitude: {round(location.latitude,2)}, Longitude: {round(location.longitude,2)}")
                        else:
                            st.error(f"❌ {get_text('Location not found in Cameroon.')}")
                    if st.form_submit_button(get_text("verify")):
                        pass
                    Nationalite_ = st.selectbox(get_text("Country"), options=pays_afrique_centrale)
                    Age =(Date_remplissage - Date_naiss).days //365
                    Religion_ = st.selectbox(get_text("Religion"), ["Christianity", "Islam", "Judaism", "Buddhism", "Other", "No religion"])
                    Situation_Matrimoniale_SM = st.selectbox(get_text("Marital Status"), ["Single", "Married", "Divorced", "Widowed", "Domestic Partnership"])
                    if A_deja_donne_le_sang_ == "Yes":
                        Date_dernier_don_ = st.date_input(get_text("If already donated, specify the date of the last donation"))
                    else:
                        Date_dernier_don_ = None  
                    Taux_dhemoglobine_ = st.number_input(get_text("Hemoglobin Level"), min_value=0.0, max_value=25.0, step=0.1)               
                    #ELIGIBILITE_AU_DON = st.selectbox(get_text("ELIGIBILITY FOR DONATION"), ["Not Eligible", "Eligible"])
                # Create a container for non-eligibility fields that will be shown/hidden based on eligibility
                non_eligibility_container = st.container()
                
                # Use an empty element to control visibility, which will be filled if Not Eligible is selected
                with non_eligibility_container:
                    col3, col4 ,col5= st.columns(3)
                    with col3:
                            Est_sous_anti_biotherapie = st.radio(get_text("Under antibiotic therapy?"), ["No", "Yes"])
                            Taux_dhemoglobine_bas = st.radio(get_text("Low hemoglobin level?"), ["No", "Yes"])
                            Date_dernier_don_3_mois = st.radio(get_text("Last donation date < 3 months?"), ["No", "Yes"])
                            IST_recente = st.radio(get_text("Recent STI (Excluding HIV, Hbs, Hcv)?"), ["No", "Yes"])
                            DDR_incorrecte = st.radio(get_text("Incorrect DDR if < 14 days before donation?"), ["No", "Yes"])
                            Allaitement = st.radio(get_text("Breastfeeding?"), ["No", "Yes"])
                    with col4:    
                            Accouchement_6mois = st.radio(get_text("Gave birth in the last 6 months?"), ["No", "Yes"])
                            Interruption_grossesse = st.radio(get_text("Pregnancy termination in the last 6 months?"), ["No", "Yes"])
                            Enceinte = st.radio(get_text("Pregnant?"), ["No", "Yes"])
                            Antecedent_transfusion = st.radio(get_text("History of transfusion?"), ["No", "Yes"])
                            Porteur_HIV_HBS_HCV = st.radio(get_text("Carrier (HIV, Hbs, Hcv)?"), ["No", "Yes"])
                            Opere = st.radio(get_text("Operated?"), ["No", "Yes"])
                    with col5:
                            Drepanocytaire = st.radio(get_text("Sickle cell?"), ["No","Yes", ])
                            Diabetique = st.radio(get_text("Diabetic?"), [ "No","Yes"])
                            Hypertendus = st.radio(get_text("Hypertensive?"), ["No", "Yes"])
                            Asthmatiques = st.radio(get_text("Asthmatic?"), ["No", "Yes"])
                            Cardiaque = st.radio(get_text("Cardiac?"), ["No", "Yes"])
                            Tatoue = st.radio(get_text("Tattooed?"), ["No", "Yes"])
                            Scarifie = st.radio(get_text("Scarified?"), ["No", "Yes"])
                        
                    st.text_input(get_text("Other reasons, specify"))
                        #Selectionner_ok = st.radio(get_text('Select "ok" to submit'), ["Ok", "No"])
                    Si_autres_raison = st.text_input(get_text("If other reasons, specify"))
                        
                if st.form_submit_button(get_text("Submit")):
                    if not Age or not Arrondissement_de_residence_ :
                        st.error("Please fill all required fields. It seems that you don't fill the date of birth")
                    else:
                        new_donor = pd.DataFrame({
                        'id': [f"DONOR_{donor_number}"],
                        'age': [Age],
                        'Date_remplissage': [Date_remplissage.strftime('%Y-%m-%d')],
                        'Date_naiss': [Date_naiss.strftime('%Y-%m-%d')],
                        'niveau_detude': [Niveau_detude],
                        'genre': [Genre_],
                        'taille': [Taille_],
                        'poids': [Poids],
                        'situation_matrimoniale': [Situation_Matrimoniale_SM],
                        'profession': [Profession_],
                        'arrondissement_residence': [Arrondissement_de_residence_],
                        'quartier_residence': [Quartier_de_Residence_],
                        'nationalite': [Nationalite_],
                        'religion': [Religion_],
                        'deja_donne_sang': [A_deja_donne_le_sang_],
                        'date_dernier_don': [Date_dernier_don_ if A_deja_donne_le_sang_ == 'Yes' else ''],
                        'taux_dhemoglobine': [Taux_dhemoglobine_],
                        'eligibilite_au_don': [None],
                        'lattitude': [latitude],
                        'longitude': [longitude],
                        'est_sous_anti_biotherapie': [Est_sous_anti_biotherapie ],
                        'taux_dhemoglobine_bas': [Taux_dhemoglobine_bas ],
                        'date_dernier_don_3_mois': [Date_dernier_don_3_mois ],
                        'ist_recente': [IST_recente ],
                        'ddr_incorrecte': [DDR_incorrecte ],
                        'allaitement': [Allaitement ],
                        'accouchement_6mois': [Accouchement_6mois ],
                        'interruption_grossesse': [Interruption_grossesse ],
                        'enceinte': [Enceinte ],
                        'antecedent_transfusion': [Antecedent_transfusion ],
                        'porteur_hiv_hbs_hcv': [Porteur_HIV_HBS_HCV ],
                        'opere': [Opere ],
                        'drepanocytaire': [Drepanocytaire ],
                        'diabetique': [Diabetique ],
                        'hypertendus': [Hypertendus ],
                        'asthmatiques': [Asthmatiques ],
                        'cardiaque': [Cardiaque ],
                        'tatoue': [Tatoue ],
                        'scarifie': [Scarifie ],
                        'autres_raisons': [Si_autres_raison ]
                    })
                        new_donor['eligibilite_au_don'] = check_eligibility(new_donor)['eligibility']
                        new_donor['raison'] = [check_eligibility(new_donor)['reasons']]
                        st.session_state.donors = pd.concat([st.session_state.donors, new_donor], ignore_index=True)
                        st.success(f"Thank you! Donor {ID} has been successfully registered.")
    #________________________________________
                        excel_file = "donnees.xlsx"

                        # Charger le fichier Excel existant
                        book = load_workbook(excel_file)

                        # Vérifier si 'Feuil1' existe
                        if 'Feuil1' in book.sheetnames:
                            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                                # Écrire le DataFrame à partir de la première ligne vide
                                startrow = book['Feuil1'].max_row  # Trouver la première ligne vide
                                new_donor.to_excel(writer, index=False, header=False, startrow=startrow, sheet_name='Feuil1')
        #_________________________________________
                    
                    # Check eligibility and display message 
                        if check_eligibility(new_donor)['eligibility'] == 'Eligible':
                            if st.session_state.language == "English" :
                                message("🎉🎆🎊 The Volonteer is  Eligible 🎉🎆🎊")
                                time.sleep(2)  
                                st.balloons()
                            else : 
                                message("🎉🎆🎊Le volontaire est  Eligible 🎉🎆🎊")
                                time.sleep(2)  
                                st.balloons()

                        elif  check_eligibility(new_donor)['eligibility'] =='Temporairement Non-eligible':
                            if st.session_state.language == "English" :
                                message(f"""The Volonteer is  Temporarily Non Eligible  😔😔""")
                            else : 
                                message(f"""Le volontaire est  Temporairement Non-eligible😔😔""")
                        else : 
                            if st.session_state.language == "English" :
                                message(f"""The Volonteer is Definitely Non Eligible  😔😔""")
                            else : 
                                message(f"""Le volontaire est  Definitivement Non-eligible😔😔""")
        

                    
                    # Display registered donors
                    st.markdown("<h2 class='sub-header'>Recent Donors</h2>", unsafe_allow_html=True)
                    recent_donors = st.session_state.donors.tail(5).copy()
                    display_columns = ['id', 'age', 'genre', 'taille', 'poids', 'situation_matrimoniale', 'eligibilite_au_don', 'date_dernier_don', 'taux_dhemoglobine']
                    st.dataframe(recent_donors[display_columns])
        
        if dataset_button:
            st.markdown(f"<h2 class='sub-header'>{get_text('NOUVELLE BASE DE DONNEES')}</h2>", unsafe_allow_html=True)
            st.dataframe(pd.read_excel("donnees.xlsx"))

    elif selected_item == "Deep insights":
        st.title("Deep insights")
        st.write("Here you can track Deep insights")
        with st.sidebar:
            donors_button = st.sidebar.button("Donors", use_container_width=True)

    elif selected_item == "Cartography":
        # Initialize session state for button states if they don't exist
        if 'choropleth_active' not in st.session_state:
            st.session_state.choropleth_active = False
        if 'marquer_active' not in st.session_state:
            st.session_state.marquer_active = False
        if 'current_view' not in st.session_state:
            st.session_state.current_view = "none"
        
        # Button functions to toggle states
        def toggle_choropleth():
            st.session_state.choropleth_active = True
            st.session_state.marquer_active = False
            st.session_state.current_view = "choropleth"
        
        def toggle_marquer():
            st.session_state.marquer_active = True
            st.session_state.choropleth_active = False
            st.session_state.current_view = "marquer"
        
        with st.sidebar:
            choropleth_button = st.sidebar.button(get_text("Choropleth"), key="Choropleth", on_click=toggle_choropleth, use_container_width=True)
            marquer_button = st.sidebar.button(get_text("Marquers"), key="Marquers", on_click=toggle_marquer, use_container_width=True)
        
        st.markdown("""
            <style>
            .block-container {
                padding-top: 1.9rem;
                padding-bottom: 0rem;
                padding-left: 1rem;
                padding-right: 2rem;
            }
            </style>
            """, unsafe_allow_html=True)
        

        # Affichage de la carte choroplèthe
        if st.session_state.choropleth_active or not st.session_state.marquer_active:
            gdf = load_shapefile("gadm41_CMR_0.shp")
            m = folium.Map(location=[7.87, 11.52], zoom_start=6, control_scale=True, tiles="CartoDB positron", height=901)
            folium.GeoJson(
                gdf.to_crs(epsg=4326).__geo_interface__,
                name="Cameroun",
                style_function=lambda x: {'fillColor': 'transparent', 'color': 'black', 'weight': 2, 'fillOpacity': 0.0}
            ).add_to(m)
            m.fit_bounds([[4.2, 7.8], [8, 17.75]])
            col1, col2 = st.columns([3.5, 1.5])
            # Créer la carte de base seulement si nécessaire
            with col2:
                st.markdown(f"<h2 class='card-title' style='text-align: center; color: #8a2be2;'>{get_text('Some Filters')}</h2>", unsafe_allow_html=True)
                st.markdown(f"<h2  style='text-align: center; font-size: 16px;'>{get_text('Select a level to watch choropleth map by region, department or district')}</h2>", unsafe_allow_html=True)
                    
                st.markdown("""
                <style>
                div[data-testid="stButton"] button {
                    color: black !important;
                }
                </style>
                """, unsafe_allow_html=True)
            
                if st.button(get_text("By region"), use_container_width=True, key="region_btn")   :
                    # Charger les données prétraitées
                    data_df_1 = get_preprocessed_data(1)
                    st.markdown(f"<h2  style='text-align: center; font-size: 16px;'>{get_text('Top Regions with the highest number of volunteers')}</h2>", unsafe_allow_html=True)
                    st.dataframe(data_df_1, use_container_width=True)
                    
                    # Charger et simplifier les shapefile régionaux
                    geo_data_1 = load_shapefile("gadm41_CMR_1.shp", simplify_tolerance=0.01)
                    geo_data_1 = geo_data_1.merge(data_df_1, on='NAME_1')
                    
                    # Créer une nouvelle carte pour cette vue
                    m = folium.Map(location=[7.87, 11.52], zoom_start=6, control_scale=True, 
                                    tiles="CartoDB positron", height=901)
                    m.fit_bounds([[4.2, 7.8], [8, 17.75]])
                    
                    # Ajouter les contours nationaux
                    gdf = load_shapefile("gadm41_CMR_0.shp", simplify_tolerance=0.01)
                    folium.GeoJson(
                        gdf.to_crs(epsg=4326).__geo_interface__,
                        name="Cameroun",
                        style_function=lambda x: {'fillColor': 'transparent', 'color': 'black', 'weight': 2, 'fillOpacity': 0.0}
                    ).add_to(m)
                    
                    # Ajouter le choropleth
                    Choropleth(
                        geo_data=geo_data_1,
                        data=geo_data_1,
                        columns=['NAME_1', 'Nb'],
                        key_on='feature.properties.NAME_1',
                        fill_color='Reds',
                        fill_opacity=1,
                        line_opacity=0.6,
                    ).add_to(m)
                    
                    # Ajouter les tooltips
                    folium.GeoJson(
                        geo_data_1,
                        style_function=lambda x: {'fillOpacity': 0, 'weight': 0},
                        tooltip=folium.features.GeoJsonTooltip(
                            fields=['NAME_1', 'Nb'],
                            aliases=[get_text('Region:'), 'Nb'],
                            style=("background-color: white; color: #333333; font-family: arial; font-size: 12px; padding: 10px;")
                        )
                    ).add_to(m)
                
                if st.button(get_text("By Department"), use_container_width=True, key="dept_btn"):
                    # Charger les données prétraitées
                    data_df_2 = get_preprocessed_data(2)
                    st.markdown(f"<h2  style='text-align: center; font-size: 16px;'>{get_text('Top Departments with the highest number of volunteers')}</h2>", unsafe_allow_html=True)
                    
                    st.dataframe(data_df_2.head(10), use_container_width=True)
                    
                    # Charger et simplifier les shapefile départementaux
                    geo_data_2 = load_shapefile("gadm41_CMR_2.shp", simplify_tolerance=0.001)
                    geo_data_2 = geo_data_2.merge(data_df_2, on='NAME_2')
                    
                    # Créer une nouvelle carte pour cette vue
                    m = folium.Map(location=[7.87, 11.52], zoom_start=6, control_scale=True, 
                                tiles="CartoDB positron", height=901)
                    m.fit_bounds([[3.2, 8.8], [9, 17.4]])
                    # Ajouter les contours nationaux
                    gdf = load_shapefile("gadm41_CMR_0.shp", simplify_tolerance=0.01)
                    folium.GeoJson(
                        gdf.to_crs(epsg=4326).__geo_interface__,
                        name="Cameroun",
                        style_function=lambda x: {'fillColor': 'transparent', 'color': 'black', 'weight': 2, 'fillOpacity': 0.0}
                    ).add_to(m)
                    
                    # Ajouter le choropleth
                    Choropleth(
                        geo_data=geo_data_2,
                        data=geo_data_2,
                        columns=['NAME_2', 'Nb'],
                        key_on='feature.properties.NAME_2',
                        fill_color='Reds',
                        fill_opacity=1,
                        line_opacity=0.6,
                        highlight=True
                    ).add_to(m)
                    
                    # Ajouter les tooltips
                    folium.GeoJson(
                        geo_data_2,
                        style_function=lambda x: {'fillOpacity': 0, 'weight': 0},
                        tooltip=folium.features.GeoJsonTooltip(
                            fields=['NAME_2','Nb'],
                            aliases=[get_text('Départment:'), 'Nb'],
                            style=("background-color: white; color: #333333; font-family: arial; font-size: 12px; padding: 10px;")
                        )
                    ).add_to(m)    
                if st.button(get_text("By District"), use_container_width=True, key="dist_btn"):
                    # Charger les données prétraitées
                    data_df_3 = get_preprocessed_data(3)
                    st.markdown(f"<h2  style='text-align: center; font-size: 16px;'>{get_text('Top Districts with the highest number of volunteers')}</h2>", unsafe_allow_html=True)
                    st.dataframe(data_df_3.head(100), use_container_width=True)
                    
                    # Charger et simplifier les shapefiles des arrondissements
                    geo_data_3 = load_shapefile("gadm41_CMR_3.shp", simplify_tolerance=0.01)
                    geo_data_3 = geo_data_3.merge(data_df_3, left_on='NAME_3', right_on='Arrondissement_de_résidence_')
                    
                    # Créer une nouvelle carte pour cette vue
                    m = folium.Map(location=[7.87, 11.52], zoom_start=6, control_scale=True, 
                                tiles="CartoDB positron", height=901)
                    m.fit_bounds([[3.2, 8.8], [9, 17.4]])
                    # Ajouter les contours nationaux
                    gdf = load_shapefile("gadm41_CMR_0.shp", simplify_tolerance=0.01)
                    folium.GeoJson(
                        gdf.to_crs(epsg=4326).__geo_interface__,
                        name="Cameroun",
                        style_function=lambda x: {'fillColor': 'transparent', 'color': 'black', 'weight': 2, 'fillOpacity': 0.0}
                    ).add_to(m)
                    
                    # Ajouter le choropleth
                    Choropleth(
                        geo_data=geo_data_3,
                        data=geo_data_3,
                        columns=['NAME_3', 'Nb'],
                        key_on='feature.properties.NAME_3',
                        fill_color='Reds',
                        fill_opacity=1,
                        line_opacity=0.6,
                    ).add_to(m)
                    
                    # Ajouter les tooltips
                    folium.GeoJson(
                        geo_data_3,
                        style_function=lambda x: {'fillOpacity': 0, 'weight': 0},
                        tooltip=folium.features.GeoJsonTooltip(
                            fields=['NAME_3', 'Nb'],
                            aliases=[get_text('District:'), get_text('Volonteers')],
                            style=("background-color: white; color:rgb(15, 15, 15); font-family: arial; font-size: 12px; padding: 10px;")
                        )
                    ).add_to(m)
                    
                
            with col1:
                    with stylable_container(
                        key='1739',
                        css_styles=f"""
                            {{
                                width: {'100%'};   
                                border: 1px solid #c0c0c0;
                                border-radius: 10px;
                                margin: 0px;  # Small margin for spacing
                                padding: 0px;
                                display: flex;
                                flex-direction: column;
                                align-items: center;
                                justify-content: center;
                                background-color: #f8f9fa;
                                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.2);
                                transition: all 0.2s ease;
                                overflow: hidden;
                                text-overflow: ellipsis;
                                color: green;
                            }}
                            
                            .card-title {{
                                    font-weight: bold;
                                    margin: 0px;
                                    padding: 0px;
                                    font-size: 1em;
                                    text-align: center;
                                    color: #8a2be2;  # Light purple color
                                }}
                        """
                    ): folium_static(m, width=900, height=900)
        # Affichage de la carte de marqueurs
        if st.session_state.marquer_active:
            # Charger les données une seule fois
            combined_data = get_hierarchical_data()
            #combined_data = pd.read_excel('Challenge dataset.xlsx')
            #combined_data.rename(columns={'Quartier de_Résidence': 'Arrondissement_de_résidence_'}, inplace=True)
            

            # Charger et simplifier le shapefile une seule fois
            geo_data_3 = load_shapefile("gadm41_CMR_3.shp", simplify_tolerance=0.01)

            # Calculer tous les centroids en une seule fois avec une approche vectorisée
            geo_data_3['centroid_lat'] = geo_data_3['geometry'].centroid.y
            geo_data_3['centroid_lon'] = geo_data_3['geometry'].centroid.x

            # Créer un dictionnaire de centroids plus simplement
            centroid_dict = dict(zip(geo_data_3['NAME_3'], zip(geo_data_3['centroid_lat'], geo_data_3['centroid_lon'])))

            # Fusionner seulement après avoir préparé les centroids
            geo_data_3 = geo_data_3.merge(combined_data, left_on='NAME_3', right_on='Arrondissement_de_résidence_')
            # Identifier les indices des lignes avec coordonnées manquantes
            mask_missing = combined_data['latitude'].isna() | combined_data['longitude'].isna()
            missing_indices = combined_data[mask_missing].index

            # Appliquer les centroids en une seule fois pour chaque colonne
            for idx in missing_indices:
                arrondissement = combined_data.loc[idx, 'Arrondissement_de_résidence_']
                if arrondissement in centroid_dict:
                    combined_data.loc[idx, 'latitude'] = centroid_dict[arrondissement][0]
                    combined_data.loc[idx, 'longitude'] = centroid_dict[arrondissement][1]
    #___________________________________________
            combined_data = combined_data.dropna(subset=['latitude', 'longitude'])
            
            # Créer la carte de base
            if 'marker_map' not in st.session_state or st.session_state.current_view != "marquer":
                gdf = load_shapefile("gadm41_CMR_0.shp", simplify_tolerance=0.01)
                m1 = folium.Map(location=[7.87, 11.52], zoom_start=4, control_scale=True, 
                    tiles="CartoDB positron", height=901)

                folium.GeoJson(
                    gdf.to_crs(epsg=4326).__geo_interface__,
                    name="Cameroun",
                    style_function=lambda x: {'fillColor': 'transparent', 'color': 'black', 'weight': 2, 'fillOpacity': 0.0}
                ).add_to(m1) 
                st.session_state.marker_map = m1
            
            # Colonnes d'affichage
            col1, col2 = st.columns([3.5, 1.5])
            
            with col2:
                st.markdown(f"<h2 class='card-title' style='text-align: center; color: #8a2be2;'>{get_text('Some Filters')}</h2>", unsafe_allow_html=True)
                st.markdown(f"<h7 class='card-title' style='text-align: center; '>{get_text('By default, Nothing selected means all is selected')}</h7>", unsafe_allow_html=True)
                
                all_arrondissement = combined_data['Arrondissement_de_résidence_'].unique()
                selected_arrondissement = st.multiselect(
                    get_text("Districts"),
                    all_arrondissement
                )
                if not selected_arrondissement:
                    selected_arrondissement = all_arrondissement

                # Filtres optimisés
                all_marital = sorted(combined_data['Situation_Matrimoniale_(SM)'].unique())
                selected_marital = st.multiselect(
                    "Marital status",
                    all_marital,  # Limiter par défaut pour améliorer les performances
                )

                if not selected_marital:
                    selected_marital = all_marital

                all_case = combined_data['ÉLIGIBILITÉ_AU_DON.'].unique()
                selected_case = st.multiselect(
                    "Eligible",
                    all_case # Limiter par défaut pour améliorer les performances
                )
                if not selected_case:
                    selected_case = all_case

                all_gender = sorted(combined_data['Genre_'].unique())
                selected_gender = st.multiselect(
                    get_text("Gender"),
                    all_gender # Limiter par défaut pour améliorer les performances
                )
                if not selected_gender:
                    selected_gender = all_gender

                # min_age = int(combined_data['Age'].min())
                # max_age = int(combined_data['Age'].max())
                # age_range = st.slider(
                #     get_text("Select age range"),
                #     min_value=min_age,
                #     max_value=max_age,
                #     value=(min_age, max_age),
                #     step=1
                # )

                all_level = sorted(combined_data["Niveau_d'etude"].unique())
                selected_level = st.multiselect(
                    get_text("Level"),
                    all_level # Limiter par défaut pour améliorer les performances
                )
                if not selected_level:
                    selected_level = all_level
                
                # Filtrer les données avec tous les filtres
                filtered_data = combined_data[
                    (combined_data['Arrondissement_de_résidence_'].isin(selected_arrondissement)) &
                    (combined_data['Situation_Matrimoniale_(SM)'].isin(selected_marital)) &
                    (combined_data['ÉLIGIBILITÉ_AU_DON.'].isin(selected_case)) &
                    (combined_data['Genre_'].isin(selected_gender)) &
                    #(combined_data['Age'] >= age_range[0]) &
                    #(combined_data['Age'] <= age_range[1]) &
                    (combined_data["Niveau_d'etude"].isin(selected_level))
                    
    ]

                # Grouper par quartier pour optimiser le rendu des marqueurs
                quartier_counts = filtered_data['Quartier_de_Résidence_'].value_counts().reset_index()
                quartier_counts.columns = ['Quartier_de_Résidence_', 'count']
                
                # Fusionner avec les coordonnées
                quartier_locations = filtered_data.groupby('Quartier_de_Résidence_').agg({
                    'latitude': 'first',
                    'longitude': 'first'
                }).reset_index()
                
                quartier_data = quartier_counts.merge(quartier_locations, on='Quartier_de_Résidence_')
                
                # Déterminer les tailles des marqueurs
                min_count = quartier_data['count'].min() if not quartier_data.empty else 1
                max_count = quartier_data['count'].max() if not quartier_data.empty else 1
                
                # Créer une nouvelle carte
                m1 = folium.Map(location=[7.87, 11], zoom_start=6, control_scale=True,height=901)

                folium.TileLayer(
                    tiles="CartoDB positron",
                    attr="CartoDB",
                    name="CartoDB Positron",
                    overlay=False,
                    control=True
                ).add_to(m1)
                
                # Ajouter les contours nationaux
                gdf = load_shapefile("gadm41_CMR_0.shp", simplify_tolerance=0.01)
                folium.GeoJson(
                    gdf.to_crs(epsg=4326).__geo_interface__,
                    name="Cameroun",
                    style_function=lambda x: {'fillColor': 'transparent', 'color': 'black', 'weight': 2, 'fillOpacity': 0.0}
                ).add_to(m1)
                m1.fit_bounds([[4.2, 7.8], [8, 17.75]])
                
                # Ajouter les marqueurs de manière optimisée
                for _, row in quartier_data.iterrows():
                    quartier = row["Quartier_de_Résidence_"]
                    count = row["count"]
                    
                    # Calculer la taille du marqueur
                    radius =  3+4* ((count - min_count) / max(max_count - min_count, 1)) * 10
                    
                    folium.CircleMarker(
                        location=[row["latitude"], row["longitude"]],
                        radius=radius,
                        tooltip=f"{quartier}: {count} {get_text('Volonteers')}",
                        color='red',
                        fill=True,
                        fill_color='red'
                    ).add_to(m1)
                
                # Ajouter contrôle de couches
                folium.LayerControl().add_to(m1)
                
            # Afficher la carte
            with col1:
                st.markdown(f"<h3 class='card-title' style='text-align: center; ;'>{get_text('Zoom for better appreciation')}</h3>", unsafe_allow_html=True)
                with stylable_container(
                    key='1946',
                    css_styles=f"""
                        {{
                            width: {'100%'};   
                            border: 1px solid #c0c0c0;
                            border-radius: 10px;
                            margin: 0px;
                            padding: 0px;
                            display: flex;
                            flex-direction: column;
                            align-items: center;
                            justify-content: center;
                            background-color: #f8f9fa;
                            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.2);
                            transition: all 0.2s ease;
                            overflow: hidden;
                            text-overflow: ellipsis;
                            color: green;
                        }}

                    """
                ): folium_static(m1, width=900, height=900)              

    elif selected_item == "About":
        with st.sidebar:
            Blood_Donation = st.sidebar.button(get_text("Blood Donation"), use_container_width=True)
            contact_us  = st.sidebar.button(get_text("Contact Us"), use_container_width=True)

        if Blood_Donation:
            st.markdown(f"<h1 class='main-header'>{get_text('About Blood Donation')}</h1>", unsafe_allow_html=True)
            
            st.markdown(f"""
            <h2 class='sub-header'>{get_text('Why Donate Blood?')}</h2>
            <p class='info-text'>{get_text('Blood donation is a critical lifesaving process that helps millions of people every year. A single donation can save up to three lives, and someone needs blood every two seconds.')}</p>
            
            <div class='highlight'>
            <h3>{get_text('Benefits of Donating Blood:')}</h3>
            <ul>
                <li>{get_text('Helps save lives')}</li>
                <li>{get_text('Free health screening')}</li>
                <li>{get_text('Reduces risk of heart disease')}</li>
                <li>{get_text('Reduces risk of cancer')}</li>
                <li>{get_text('Helps in weight loss')}</li>
                <li>{get_text('Helps in replenishing blood cells')}</li>
            </ul>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown(f"""
                <h2 class='sub-header'>{get_text('Eligibility Requirements')}</h2>
                <ul>
                    <li>{get_text('Age: 18-65 years')}</li>
                    <li>{get_text('Weight: At least 50 kg')}</li>
                    <li>{get_text('Hemoglobin: 12.5 g/dL for women, 13.0 g/dL for men')}</li>
                    <li>{get_text('Good general health')}</li>
                    <li>{get_text('No fever or active infection')}</li>
                    <li>{get_text('No recent tattoos or piercings (within 4 months)')}</li>
                    <li>{get_text('No recent major surgery')}</li>
                    <li>{get_text('No high-risk behaviors')}</li>
                </ul>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                    <h2 class='sub-header'>{get_text('The Donation Process')}</h2>
                    <ol>
                        <li><strong>{get_text('Registration:')}</strong> {get_text('Complete a donor registration form')}</li>
                        <li><strong>{get_text('Health History:')}</strong> {get_text('Answer questions about your health history')}</li>
                        <li><strong>{get_text('Mini-Physical:')}</strong> {get_text('Check temperature, pulse, blood pressure, and hemoglobin')}</li>
                        <li><strong>{get_text('Donation:')}</strong> {get_text('The actual donation takes about 8-10 minutes')}</li>
                        <li><strong>{get_text('Refreshments:')}</strong> {get_text('Rest and enjoy refreshments for 15 minutes')}</li>
                    </ol>
                """, unsafe_allow_html=True)
            
            st.markdown(f"""
            <h2 class='sub-header'>{get_text('Frequently Asked Questions')}</h2>
            """, unsafe_allow_html=True)
            
            faq_expander = st.expander(get_text("Click to view FAQs"))
            
            st.markdown(f"""
            <h2 class='sub-header'>{get_text('Blood Donation Facts')}</h2>
            <div class='highlight'>
            <ul>
                <li>{get_text('Every 2 seconds someone needs blood')}</li>
                <li>{get_text('A single car accident victim can require up to 100 units of blood')}</li>
                <li>{get_text('One donation can save up to 3 lives')}</li>
                <li>{get_text('Only 37% of the population is eligible to donate blood')}</li>
                <li>{get_text('Less than 10% of eligible donors actually donate')}</li>
                <li>{get_text('Blood cannot be manufactured – it can only come from donors')}</li>
                <li>{get_text('Red blood cells have a shelf life of 42 days')}</li>
                <li>{get_text('Platelets have a shelf life of just 5 days')}</li>
            </ul>
            </div>
            """, unsafe_allow_html=True)
        
        if contact_us:
            # About Page Title
            st.markdown(
                f"""
                <h1 style='color: red; text-align: center;'>{get_text('About Our Team')}</h1>
                """,
                unsafe_allow_html=True
            )
            
            # Introduction Text
            st.markdown(
                get_text("Welcome to the About section! Our team is composed of passionate and skilled engineering students in statistics and economics from the Institut Sous-Régional de Statistique et d Économie Appliquée (ISSEA). We are dedicated to building insightful and interactive data-driven solutions, leveraging our expertise in statistical modeling, data visualisations, and economic analysis.") 
            )

            # Display Team Members in a 2x2 Grid
            col1, col2 = st.columns(2)
            @st.cache_resource
            def get_image_path(image_name):
                if image_name.startswith("http"):
                    return image_name  # Pas besoin de modifier les URLs externes
                else:
                    # Si c'est un chemin local, ajuster pour qu'il soit relatif au dossier img
                    # Enlever le "/" initial si présent
                    if image_name.startswith("/"):
                        image_name = image_name[1:]
                    return image_name

            team_members = [
                {"name": "ASSA ALLO", "email": "alloassa21@gmail.com", "phone": "+1237 _________", "image": "img/aa.jpg"},
                {"name": " TAKOUGOUM Steeve Rodrigue", "email": "rodriguetakougoum@gmail.com", "phone": "+237__________", "image": "img/sr.png"},
                {"name": " TIDJANI Razak", "email": "tidjanirazak0@gmail.com", "phone": "+237 ___________", "image": "img/rz.jpg"},            
                {"name": "TCHINDA Chris Donald", "email": "tcd9602@gmail.com", "phone": "+237 ___________ ", "image": "img/tcd.jpg"},
            ]

            # Display Team Members in a 1x4 Grid
            cols = st.columns(4)
            for i, member in enumerate(team_members):
                with cols[i % 4]:
                    image_path = get_image_path(member['image'])
                    
                    st.image(image_path, width=180)
                    
                    st.markdown(
                        f"""
                        <h6>{member['name']}</h6>
                        <p style="font-size: 13px;">✉️ {member['email']}</p>
                        <p style="font-size: 11px;">📞 {member['phone']}</p>
                        """,
                        unsafe_allow_html=True
                    )
                    ""


def main():
    # Si l'utilisateur n'est pas authentifié, afficher le formulaire de connexion
    if not st.session_state["authenticated"]:
        login_form()
    else:
        # Si l'utilisateur est authentifié, afficher le contenu principal
        main_app()

if __name__ == "__main__":
    main()


